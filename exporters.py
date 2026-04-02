"""Excel and PDF exporters for parsed EML reports."""

import logging
import os
import sys
from pathlib import Path
from typing import Any, Dict, List

from models import ParsedReport

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Excel exporter (openpyxl)
# ---------------------------------------------------------------------------

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_AVAILABLE = True
except ImportError:
    _OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl није инсталиран; Excel извоз неће бити доступан.")


class ExcelExporter:
    """Export a :class:`ParsedReport` to an .xlsx file using openpyxl."""

    # Dark header style colours
    _HEADER_BG = "2D6A9F"
    _HEADER_FG = "FFFFFF"
    _ALT_ROW_BG = "EAF2FB"

    def export(self, report: ParsedReport, path: str | Path) -> None:
        """
        Write the report to *path* as an Excel workbook.

        Raises:
            RuntimeError: If openpyxl is not installed.
            OSError: If the file cannot be written.
        """
        if not _OPENPYXL_AVAILABLE:
            raise RuntimeError(
                "Библиотека openpyxl није инсталирана. "
                "Покрените: pip install openpyxl"
            )

        wb = openpyxl.Workbook()

        self._write_overview_sheet(wb, report)
        self._write_workspace_sheet(wb, "Активни", report.active_workspaces)
        self._write_workspace_sheet(wb, "Старији од 90 дана", report.old_workspaces)

        # Remove the default empty sheet created by openpyxl
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

        wb.save(str(path))
        logger.info("Excel датотека сачувана: %s", path)

    # ------------------------------------------------------------------
    # Sheet builders
    # ------------------------------------------------------------------

    def _write_overview_sheet(self, wb: Any, report: ParsedReport) -> None:
        ws = wb.create_sheet("Преглед")

        header_font = Font(bold=True, color=self._HEADER_FG)
        header_fill = PatternFill("solid", fgColor=self._HEADER_BG)

        def _hdr(cell: Any) -> None:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="left", vertical="center")

        rows: List[tuple] = [
            ("Поље", "Вредност"),
            ("Пошиљалац", report.from_addr),
            ("Прималац", report.to_addr),
            ("Датум", report.date),
            ("Наслов", report.subject),
            (),  # blank separator
            ("Волумен", report.volume_name),
            ("Слободан простор", report.free_space),
            ("Искоришћено %", report.used_percent),
            ("Промена %", report.change_percent),
            (),  # blank separator
            ("Број активних", len(report.active_workspaces)),
            ("Број старих (> 90 дана)", len(report.old_workspaces)),
        ]

        for i, row in enumerate(rows, start=1):
            ws.append(list(row))
            if i == 1:
                _hdr(ws.cell(i, 1))
                _hdr(ws.cell(i, 2))

        # Freeze header row
        ws.freeze_panes = "A2"

        self._autofit(ws, [30, 50])

    def _write_workspace_sheet(
        self, wb: Any, title: str, workspaces: List[str]
    ) -> None:
        ws = wb.create_sheet(title)

        header_font = Font(bold=True, color=self._HEADER_FG)
        header_fill = PatternFill("solid", fgColor=self._HEADER_BG)

        ws.append(["#", "Радни простор"])
        ws.cell(1, 1).font = header_font
        ws.cell(1, 1).fill = header_fill
        ws.cell(1, 2).font = header_font
        ws.cell(1, 2).fill = header_fill

        alt_fill = PatternFill("solid", fgColor=self._ALT_ROW_BG)
        for idx, name in enumerate(workspaces, start=1):
            ws.append([idx, name])
            if idx % 2 == 0:
                ws.cell(idx + 1, 1).fill = alt_fill
                ws.cell(idx + 1, 2).fill = alt_fill

        ws.auto_filter.ref = f"A1:B{len(workspaces) + 1}"
        ws.freeze_panes = "A2"
        self._autofit(ws, [6, 60])

    @staticmethod
    def _autofit(ws: Any, widths: List[int]) -> None:
        """Apply fixed column widths (openpyxl cannot measure actual text width)."""
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------------------
# PDF exporter (reportlab)
# ---------------------------------------------------------------------------

try:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.platypus import (
        Paragraph,
        SimpleDocTemplate,
        Spacer,
        Table,
        TableStyle,
        PageBreak,
    )
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    _REPORTLAB_AVAILABLE = True
except ImportError:
    _REPORTLAB_AVAILABLE = False
    logger.warning("reportlab није инсталиран; PDF извоз неће бити доступан.")


def _register_cyrillic_font() -> str:
    """
    Register a font that supports Cyrillic.

    Tries, in order:
    1. DejaVuSans shipped by reportlab (if present)
    2. A system DejaVuSans.ttf
    3. Falls back to Helvetica (Latin only, will show boxes for Cyrillic).

    Returns the registered font name.
    """
    if not _REPORTLAB_AVAILABLE:
        return "Helvetica"

    font_name = "DejaVuSans"
    font_name_bold = "DejaVuSans-Bold"

    # Common font search paths
    search_dirs = [
        "/usr/share/fonts",
        "/usr/local/share/fonts",
        os.path.expanduser("~/.fonts"),
        os.path.expanduser("~/Library/Fonts"),         # macOS
        "C:/Windows/Fonts",                             # Windows
        os.path.join(sys.prefix, "share", "fonts"),
    ]

    def _find_file(name: str) -> str:
        for d in search_dirs:
            for root, _dirs, files in os.walk(d):
                if name in files:
                    return os.path.join(root, name)
        return ""

    regular = _find_file("DejaVuSans.ttf")
    bold = _find_file("DejaVuSans-Bold.ttf")

    if regular:
        try:
            pdfmetrics.registerFont(TTFont(font_name, regular))
            if bold:
                pdfmetrics.registerFont(TTFont(font_name_bold, bold))
            else:
                font_name_bold = font_name  # use regular as bold fallback
            return font_name
        except Exception as exc:
            logger.warning("Не могу учитати DejaVuSans: %s", exc)

    logger.warning(
        "DejaVuSans.ttf није пронађен; ћирилица у PDF-у можда неће бити "
        "приказана исправно."
    )
    return "Helvetica"


class PdfExporter:
    """Export a :class:`ParsedReport` to a PDF file using reportlab."""

    def export(self, report: ParsedReport, path: str | Path) -> None:
        """
        Write the report to *path* as a PDF.

        Raises:
            RuntimeError: If reportlab is not installed.
            OSError: If the file cannot be written.
        """
        if not _REPORTLAB_AVAILABLE:
            raise RuntimeError(
                "Библиотека reportlab није инсталирана. "
                "Покрените: pip install reportlab"
            )

        font_name = _register_cyrillic_font()
        font_bold = font_name + "-Bold" if font_name == "DejaVuSans" else font_name

        styles = getSampleStyleSheet()

        title_style = ParagraphStyle(
            "CustomTitle",
            fontName=font_bold,
            fontSize=16,
            spaceAfter=12,
            textColor=colors.HexColor("#1A3C5E"),
        )
        h2_style = ParagraphStyle(
            "CustomH2",
            fontName=font_bold,
            fontSize=12,
            spaceBefore=12,
            spaceAfter=6,
            textColor=colors.HexColor("#2D6A9F"),
        )
        body_style = ParagraphStyle(
            "CustomBody",
            fontName=font_name,
            fontSize=10,
            leading=14,
        )
        item_style = ParagraphStyle(
            "CustomItem",
            fontName=font_name,
            fontSize=9,
            leading=13,
            leftIndent=12,
        )

        doc = SimpleDocTemplate(
            str(path),
            pagesize=A4,
            rightMargin=2 * cm,
            leftMargin=2 * cm,
            topMargin=2.5 * cm,
            bottomMargin=2 * cm,
        )

        story = []

        # Title
        story.append(Paragraph("ELEMENTS Info: Извештај о статусу пројеката", title_style))
        story.append(Spacer(1, 0.4 * cm))

        # Metadata table
        story.append(Paragraph("Метаподаци поруке", h2_style))
        meta_data = [
            ["Пошиљалац", report.from_addr or "—"],
            ["Прималац", report.to_addr or "—"],
            ["Датум", report.date or "—"],
            ["Наслов", report.subject or "—"],
        ]
        story.append(self._make_table(meta_data, font_name, font_bold, col_widths=[4 * cm, 13 * cm]))

        story.append(Spacer(1, 0.4 * cm))

        # Storage table
        story.append(Paragraph("Коришћење складишта", h2_style))
        storage_data = [
            ["Волумен", report.volume_name or "—"],
            ["Слободан простор", report.free_space or "—"],
            ["Искоришћено", report.used_percent or "—"],
            ["Промена", report.change_percent or "—"],
        ]
        story.append(self._make_table(storage_data, font_name, font_bold, col_widths=[4 * cm, 13 * cm]))

        story.append(Spacer(1, 0.4 * cm))

        # Summary counters
        story.append(Paragraph("Сажетак", h2_style))
        summary_data = [
            ["Број активних радних простора", str(len(report.active_workspaces))],
            ["Број старих радних простора (> 90 дана)", str(len(report.old_workspaces))],
        ]
        story.append(self._make_table(summary_data, font_name, font_bold, col_widths=[10 * cm, 7 * cm]))

        story.append(Spacer(1, 0.6 * cm))

        # Active workspaces
        story.append(Paragraph("Активни радни простори", h2_style))
        if report.active_workspaces:
            for ws_name in report.active_workspaces:
                story.append(Paragraph(f"• {ws_name}", item_style))
        else:
            story.append(Paragraph("(нема активних радних простора)", body_style))

        story.append(Spacer(1, 0.6 * cm))

        # Old workspaces
        story.append(Paragraph("Радни простори старији од 90 дана", h2_style))
        if report.old_workspaces:
            for ws_name in report.old_workspaces:
                story.append(Paragraph(f"• {ws_name}", item_style))
        else:
            story.append(Paragraph("(нема старих радних простора)", body_style))

        doc.build(story)
        logger.info("PDF датотека сачувана: %s", path)

    @staticmethod
    def _make_table(
        data: List[List[str]],
        font_name: str,
        font_bold: str,
        col_widths: List[float],
    ) -> "Table":
        """Build a two-column key/value table with consistent styling."""
        from reportlab.lib import colors as _colors
        from reportlab.platypus import Table as _Table, TableStyle as _TableStyle

        tbl = _Table(data, colWidths=col_widths)
        tbl.setStyle(
            _TableStyle(
                [
                    ("FONTNAME", (0, 0), (0, -1), font_bold),
                    ("FONTNAME", (1, 0), (1, -1), font_name),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("BACKGROUND", (0, 0), (0, -1), _colors.HexColor("#EAF2FB")),
                    ("GRID", (0, 0), (-1, -1), 0.5, _colors.HexColor("#CCCCCC")),
                    ("ROWBACKGROUNDS", (0, 0), (-1, -1),
                     [_colors.HexColor("#FFFFFF"), _colors.HexColor("#F5F9FE")]),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return tbl
